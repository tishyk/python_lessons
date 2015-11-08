#! usr/bin/env
#! coding: utf-8

import openpyxl
import os
import time

filename = 'example_write.xlsx' # path to file
wb = openpyxl.Workbook()

wb.create_sheet(index=0,title='myReport')

wb.create_sheet(index=1,title='myReport~temp')
#and remove previously created sheet
wb.remove_sheet(wb.get_sheet_by_name('myReport~temp'))

