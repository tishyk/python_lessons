#! usr/bin/env
#! coding: utf-8

import openpyxl
import os
import time

filename = 'example.xlsx' # path to file
wb = openpyxl.load_workbook(filename) # openpyxl.load_workbook('example.xlsx')
wb_sheets = wb.get_sheet_names() #  get all sheets from *.xlsx file
for i in wb_sheets:
    print i
sheet1 = wb.get_active_sheet() # choose 1-st sheet for work
print "Sheet - %s activated!"%sheet1.title
sheet2 = wb.get_sheet_by_name(wb_sheets[1]) # choose 2-st sheet for work too
print "Sheet - %s activated!"%sheet2.title
print sheet1['A1'].value # use NAME of cell for get or set value
print sheet1.cell(row=1, column=2).value  # use cell for get or set value

for i in range(1, 8):
    print(i, sheet1.cell(row=i, column=2).value)

#How get highest row and column?
print sheet1.max_row    
print sheet1.max_column

from openpyxl.cell import get_column_letter

#How get NAME from cell row or column index?
print get_column_letter(sheet1.max_column)

# get all items from one column
print sheet1.columns[1]

# How to save title to active sheet?
sheet1.title = "Activated"

wb.save('report.xlsx')

# write into report.xlsx file next headers:
# №~   ITEM              CREATION TIME
#write into body info from list:
# 1 s.tischenko         'Tue Oct 27 21:57:22 2015'
# 2 Test_group3.docx    'Sat Oct 31 14:44:39 2015'

#How?
# get list of items in folder:  os.listdir('..') -> Ex. ['python_test1.png', 's.tischenko',...
# get item creation time: time.ctime(os.path.getctime('../python_test1.png'))
# check for item type -> os.path.isdir('../python_test1.png') -> True if directory (write <DIR>)
# else TYPE -> <File>


#**
# For strig formating use %
# "N%0.2i\t%.%is\t<%s>\t%s" % (lst.index(item)+1,max_name_lenmax_name_len,item,item_type,ctime"
raw_input()
