# ! usr/bin/env
# ! coding: utf-8

import openpyxl
import os
import time

from openpyxl.styles import Font, Alignment

filename = 'example_write.xlsx'  # path to file

wb = openpyxl.Workbook()

wb.create_sheet(index=0, title='myReport')
sheet = wb.get_sheet_by_name('myReport')

font_header = Font(name='Times New Roman', size=12, bold=True, color="850808")
font_i_header = Font(name='Calibri', size=11, bold=True, italic=True, color="2A0599")
font_item = Font(name='Times New Roman', size=10)

headers_alignment = Alignment(horizontal='center', vertical='bottom', text_rotation=0,
                    shrink_to_fit=False, wrap_text=False, indent=0)

col_name = ('A', 'B', 'C')
headers = ('#', 'ITEM', 'CREATED DATE')
row_count = 1

for c, h in zip(col_name, headers):
    sheet[c + str(row_count)] = h
    sheet[c + str(row_count)].font = font_header
    sheet[c + str(row_count)].alignment = headers_alignment 
row_count += 1
    # or sheet.cell(row=1,column=1).value = i
sheet.freeze_panes = 'A2'

os.chdir('..')
i_list = os.listdir('.')
i_list.sort()
i_dirs = [item for item in i_list if os.path.isdir(item)]
i_files = [item for item in i_list if os.path.isfile(item)]


sheet.merge_cells("%s%i:%s%i" % (col_name[0], row_count, col_name[-1], row_count))
sheet["A%i" % row_count] = 'DIR'
sheet["A%i" % row_count].font = font_i_header
sheet["A%i" % row_count].alignment = headers_alignment 

row_count += 1                
 
for item in i_dirs:
    for column in col_name:
        if column == "A":
            sheet["A%i" % row_count] = str(i_dirs.index(item) + 1)
            sheet["A%i" % row_count].font = font_item
        if column == "B":
            sheet["B%i" % row_count] = item
            sheet["B%i" % row_count].font = font_item    
        if column == "C":
            sheet["C%i" % row_count] = time.ctime(os.path.getctime(item))
            sheet["C%i" % row_count].font = font_item 
    row_count += 1     
         
sheet.merge_cells("%s%i:%s%i" % (col_name[0], row_count, col_name[-1], row_count))
sheet["A%i" % row_count] = 'FILE'

sheet["A%i" % row_count].font = font_i_header
sheet["A%i" % row_count].alignment = headers_alignment 
row_count += 1
              
for item in i_files:
    for column in col_name:
        if column == "A":
            sheet["A%i" % row_count] = str(i_files.index(item) + 1)
        if column == "B":
            sheet["B%i" % row_count] = item       
        if column == "C":
            sheet["C%i" % row_count] = time.ctime(os.path.getctime(item))
        sheet["%s%i" % (column, row_count)].font = font_item
    row_count += 1
    
sheet["B%i" % row_count], sheet["C%i" % row_count] = "TOTAL DIRS: ", len(i_dirs)
sheet["B%i" % row_count].alignment = Alignment(horizontal='right')
sheet["C%i" % row_count].alignment = Alignment(horizontal='left')
row_count += 1

sheet["B%i" % row_count], sheet["C%i" % row_count] = "TOTAL FILES: " , len(i_files)
sheet["B%i" % row_count].alignment = Alignment(horizontal='right')
sheet["C%i" % row_count].alignment = Alignment(horizontal='left')

sheet.column_dimensions["A"].width = 5
sheet.column_dimensions["B"].width = len(max(i_list, key=len)) + 10
sheet.column_dimensions["C"].width = len(time.ctime(os.path.getctime(item)))

wb.save('myReport.xlsx')
print("Done!")
